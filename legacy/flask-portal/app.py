# Ensure ece_orignal_updated is imported globally
import sys
import os
import threading
import time
import uuid

# Add parent directory to path before importing ece_orignal_updated
sys.path.append(os.path.dirname(os.path.dirname(__file__)))

import ece_orignal_updated
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, after_this_request, jsonify
import pandas as pd
import numpy as np
import re
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string
from werkzeug.utils import secure_filename

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
ALLOWED_EXTENSIONS = {'xlsx'}

# Default file paths (relative to project root)
DEFAULT_MAPPING_PATH = os.path.join(BASE_DIR, 'Course, CO and PO mapping Nov 2025 (2).xlsx')
DEFAULT_INDIRECT_PATH = os.path.join(BASE_DIR, 'indirect.xlsx')

# Mapping file configurations
MAPPING_FILES = {
    'CO-PO Mapping Nov 25': DEFAULT_MAPPING_PATH
}

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.secret_key = os.environ.get('SECRET_KEY', 'dev-fallback-key-change-in-production')

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

FILE_MAX_AGE_SECONDS = 1800  # 30 minutes
EVAL_CO_TABLE_TITLE = "% > max(50, Mean - 0.5*Std)"
EVAL_PO_TABLE_TITLE = "Weighted Avg using CO'  (% > Mean-0.5*Std)"
PO_PSO_HEADERS = [*(f'PO{i}' for i in range(1, 13)), 'PSO1', 'PSO2', 'PSO3']
RESULT_CONTEXTS = {}
DOWNLOAD_CONTEXTS = {}


def store_result_context(payload):
    result_id = uuid.uuid4().hex
    RESULT_CONTEXTS[result_id] = {
        'created_at': time.time(),
        'payload': payload,
    }
    return result_id


def get_result_context(result_id):
    entry = RESULT_CONTEXTS.get(result_id)
    if not entry:
        return None
    return entry.get('payload')


def store_download_context(excel_path):
    if not excel_path:
        return None
    download_id = uuid.uuid4().hex
    DOWNLOAD_CONTEXTS[download_id] = {
        'created_at': time.time(),
        'excel_path': excel_path,
    }
    return download_id


def pop_download_context(download_id):
    entry = DOWNLOAD_CONTEXTS.pop(download_id, None)
    if not entry:
        return None
    return entry.get('excel_path')


def cleanup_stale_contexts():
    now = time.time()

    stale_result_ids = [
        key for key, value in RESULT_CONTEXTS.items()
        if (now - value.get('created_at', now)) > FILE_MAX_AGE_SECONDS
    ]
    for key in stale_result_ids:
        payload = RESULT_CONTEXTS.pop(key, {}).get('payload', {})
        remove_file_if_exists(payload.get('excel_path'))

    stale_download_ids = [
        key for key, value in DOWNLOAD_CONTEXTS.items()
        if (now - value.get('created_at', now)) > FILE_MAX_AGE_SECONDS
    ]
    for key in stale_download_ids:
        excel_path = DOWNLOAD_CONTEXTS.pop(key, {}).get('excel_path')
        remove_file_if_exists(excel_path)


def clear_runtime_contexts():
    removed_result_contexts = len(RESULT_CONTEXTS)
    removed_download_contexts = len(DOWNLOAD_CONTEXTS)

    for entry in RESULT_CONTEXTS.values():
        payload = entry.get('payload', {})
        remove_file_if_exists(payload.get('excel_path'))

    for entry in DOWNLOAD_CONTEXTS.values():
        remove_file_if_exists(entry.get('excel_path'))

    RESULT_CONTEXTS.clear()
    DOWNLOAD_CONTEXTS.clear()

    return {
        'result_contexts': removed_result_contexts,
        'download_contexts': removed_download_contexts,
    }

def _cleanup_old_files():
    """Background thread: delete files in uploads/ older than FILE_MAX_AGE_SECONDS."""
    while True:
        time.sleep(300)  # check every 5 minutes
        now = time.time()
        try:
            cleanup_stale_contexts()
            for filename in os.listdir(UPLOAD_FOLDER):
                filepath = os.path.join(UPLOAD_FOLDER, filename)
                if os.path.isfile(filepath) and (now - os.path.getmtime(filepath)) > FILE_MAX_AGE_SECONDS:
                    try:
                        os.remove(filepath)
                        print(f"[cleanup] Removed stale file: {filepath}")
                    except Exception as e:
                        print(f"[cleanup] Could not remove {filepath}: {e}")
        except Exception as e:
            print(f"[cleanup] Error scanning uploads folder: {e}")

_cleanup_thread = threading.Thread(target=_cleanup_old_files, daemon=True)
_cleanup_thread.start()

# Download route must be after app is defined
@app.route('/download_results/<download_id>')
def download_results(download_id):
    excel_path = pop_download_context(download_id)
    if not excel_path or not os.path.exists(excel_path):
        flash('Results file not found.')
        return redirect(url_for('index'))

    @after_this_request
    def delete_excel(response):
        try:
            if os.path.exists(excel_path):
                os.remove(excel_path)
                print(f"[cleanup] Deleted results file after download: {excel_path}")
        except Exception as e:
            print(f"[cleanup] Could not delete results file {excel_path}: {e}")
        return response

    return send_file(excel_path, as_attachment=True, download_name=os.path.basename(excel_path))


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def remove_file_if_exists(path):
    if path and os.path.exists(path):
        try:
            os.remove(path)
        except Exception as e:
            print(f"[cleanup] Could not remove {path}: {e}")


def save_uploaded_file(uploaded_file, prefix):
    filename = secure_filename(uploaded_file.filename or 'upload.xlsx')
    unique_name = f"{prefix}_{uuid.uuid4().hex}_{filename}"
    saved_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_name)
    uploaded_file.save(saved_path)
    return saved_path


def normalize_summary_label(value):
    if value is None:
        return ''
    return re.sub(r'[^a-z0-9]+', '', str(value).strip().lower())


def is_co_summary_label(value):
    normalized = normalize_summary_label(value)
    return normalized.startswith('max50mean05std') and 'weightedavg' not in normalized


def is_po_summary_label(value):
    normalized = normalize_summary_label(value)
    return normalized.startswith('weightedavgusingco') and 'mean05std' in normalized


def coerce_numeric(value):
    if value is None:
        return None
    if isinstance(value, (int, float, np.number)):
        if pd.isna(value):
            return None
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def find_summary_anchor(worksheet, matcher):
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and matcher(cell.value):
                return cell
    return None


def find_summary_anchor_in_workbook(workbook, matcher):
    for worksheet in workbook.worksheets:
        anchor_cell = find_summary_anchor(worksheet, matcher)
        if anchor_cell is not None:
            return worksheet, anchor_cell
    return None, None


def normalize_cell_reference(cell_ref):
    if cell_ref is None:
        return ''
    normalized = str(cell_ref).strip().upper()
    if not normalized:
        return ''
    try:
        coordinate_from_string(normalized)
    except ValueError:
        raise ValueError(f'"{cell_ref}" is not a valid Excel cell ID.')
    return normalized


def find_numeric_start_column(worksheet, row_idx, start_col):
    for col_idx in range(start_col, worksheet.max_column + 1):
        if coerce_numeric(worksheet.cell(row_idx, col_idx).value) is not None:
            return col_idx
    return None


def extract_consecutive_numeric_values(worksheet, row_idx, start_col, count, context_label):
    extracted = []
    missing_cells = []

    for offset in range(count):
        col_idx = start_col + offset
        cell = worksheet.cell(row_idx, col_idx)
        cell_value = cell.value
        numeric_value = coerce_numeric(cell_value)
        if cell_value in (None, '') or numeric_value is None:
            missing_cells.append(cell.coordinate)
        extracted.append(numeric_value)

    if missing_cells:
        raise ValueError(
            f'Expected {count} values from {context_label}, but these cells were blank or non-numeric: '
            + ', '.join(missing_cells)
        )

    return extracted


def read_values_from_anchor_cell(worksheet, anchor_cell, count, context_label):
    if coerce_numeric(anchor_cell.value) is not None:
        start_col = anchor_cell.column
    else:
        start_col = find_numeric_start_column(worksheet, anchor_cell.row, anchor_cell.column + 1)

    if start_col is None:
        raise ValueError(f'Could not find a numeric starting point on row {anchor_cell.row} for {context_label}.')

    return extract_consecutive_numeric_values(
        worksheet,
        anchor_cell.row,
        start_col,
        count,
        context_label,
    )


def read_summary_values_from_output(output_path, expected_headers, matcher, cell_ref=None, fixed_count=None):
    workbook = load_workbook(output_path, data_only=True)
    value_count = fixed_count if fixed_count is not None else len(expected_headers)

    override_ref = normalize_cell_reference(cell_ref)
    if override_ref:
        errors = []
        for worksheet in workbook.worksheets:
            anchor_cell = worksheet[override_ref]
            matched_label = f'{worksheet.title}!{override_ref}'
            try:
                sequential_values = read_values_from_anchor_cell(
                    worksheet,
                    anchor_cell,
                    value_count,
                    matched_label,
                )
                values = dict(zip(expected_headers, sequential_values))
                return values, matched_label
            except ValueError as e:
                errors.append(f'{worksheet.title}: {e}')
        raise ValueError('Could not read the override cell across workbook sheets. ' + ' | '.join(errors))
    else:
        worksheet, anchor_cell = find_summary_anchor_in_workbook(workbook, matcher)
        if anchor_cell is None:
            raise ValueError(f'Could not locate the required comparison row in "{os.path.basename(output_path)}".')
        matched_label = f'{worksheet.title}: {str(anchor_cell.value).strip()}'

    sequential_values = read_values_from_anchor_cell(
        worksheet,
        anchor_cell,
        value_count,
        matched_label,
    )
    values = dict(zip(expected_headers, sequential_values))

    return values, matched_label


def build_delta_values(calculated_values, read_values, ordered_headers):
    delta = {}
    for header in ordered_headers:
        calc_val = calculated_values.get(header)
        read_val = read_values.get(header)
        if calc_val is None or read_val is None:
            delta[header] = None
        else:
            delta[header] = float(calc_val) - float(read_val)
    return delta


def build_eval_table(title, columns, calculated_values, read_values, matched_label):
    delta_values = build_delta_values(calculated_values, read_values, columns)
    return {
        'title': title,
        'matched_label': matched_label,
        'columns': columns,
        'rows': [
            {'label': 'Calculated', 'kind': 'calculated', 'values': [calculated_values.get(col) for col in columns]},
            {'label': 'Read From Output', 'kind': 'read', 'values': [read_values.get(col) for col in columns]},
            {'label': 'Delta (Calculated - Read)', 'kind': 'delta', 'values': [delta_values.get(col) for col in columns]},
        ],
    }


def build_included_rolls(course_path, selected_programmes, selected_branches):
    if not selected_programmes and not selected_branches:
        return None

    try:
        parsed = parse_student_rolls(course_path)
        included_rolls = set()
        has_branch_data = bool(parsed.get('rolls_by_branch'))

        for prog in selected_programmes:
            if has_branch_data and selected_branches:
                prog_branches = [
                    br for br, info in parsed['branches'].items()
                    if info.get('programme') == prog and br in selected_branches
                ]
                if prog_branches:
                    for branch in prog_branches:
                        included_rolls.update(parsed['rolls_by_branch'].get(branch, []))
                else:
                    prog_has_any_branches = any(
                        info.get('programme') == prog for info in parsed['branches'].values()
                    )
                    if not prog_has_any_branches:
                        included_rolls.update(parsed['rolls_by_programme'].get(prog, []))
            else:
                included_rolls.update(parsed['rolls_by_programme'].get(prog, []))

        return list(included_rolls) if included_rolls else None
    except Exception as e:
        print(f"[build_included_rolls] Error parsing student rolls for filtering: {e}")
        return None


def summarize_scope_selection(selected_programmes, selected_branches):
    if not selected_programmes and not selected_branches:
        return 'Default scope (non-MT/PhD filtering from the core processor)'

    parts = []
    if selected_programmes:
        parts.append('Programmes: ' + ', '.join(selected_programmes))
    if selected_branches:
        parts.append('Branches: ' + ', '.join(branch.split('::', 1)[-1] for branch in selected_branches))
    return ' | '.join(parts)


def build_evaluation_payload(
    course_path,
    mapping_path,
    course_title,
    compare_path,
    included_rolls=None,
    co_cell_ref=None,
    po_cell_ref=None,
    target_value=50,
):
    intermediate = ece_orignal_updated.main_process(
        course_path,
        mapping_path,
        course_title,
        included_rolls=included_rolls,
        target_value=target_value,
    )

    co_columns = list(intermediate.get('unique_COs', []))
    calculated_co = {
        co: coerce_numeric(intermediate['CO_stats']['pct_above'].get(co))
        for co in co_columns
    }

    available_po_values = intermediate.get('po_pso_attainment', {})
    po_columns = list(PO_PSO_HEADERS)
    calculated_po = {
        po: coerce_numeric(available_po_values.get(po))
        for po in po_columns
    }

    read_co, matched_co_label = read_summary_values_from_output(
        compare_path,
        co_columns,
        is_co_summary_label,
        cell_ref=co_cell_ref,
        fixed_count=len(co_columns),
    )
    read_po, matched_po_label = read_summary_values_from_output(
        compare_path,
        po_columns,
        is_po_summary_label,
        cell_ref=po_cell_ref,
        fixed_count=15,
    )

    return {
        'intermediate': intermediate,
        'co_table': build_eval_table(EVAL_CO_TABLE_TITLE, co_columns, calculated_co, read_co, matched_co_label),
        'po_table': build_eval_table(EVAL_PO_TABLE_TITLE, po_columns, calculated_po, read_po, matched_po_label),
    }


def get_display_filename(uploaded_file, fallback_path=''):
    if uploaded_file and getattr(uploaded_file, 'filename', None):
        return uploaded_file.filename
    return os.path.basename(fallback_path) if fallback_path else ''


def prepare_results_payload(
    course_path,
    mapping_path,
    course_title,
    included_rolls=None,
    indirect_attainment=None,
    course_filename=None,
    mapping_filename=None,
    target_value=50,
):
    indirect_attainment = indirect_attainment or {}
    intermediate = ece_orignal_updated.main_process(
        course_path,
        mapping_path,
        course_title,
        included_rolls=included_rolls,
        target_value=target_value,
    )

    excel_path = intermediate.get('excel_path', '')
    co_po_mapping_df = pd.DataFrame.from_dict(intermediate['CO_PO_mapping'], orient='index')

    direct_po_pso_attainment = pd.Series(intermediate['po_pso_attainment'])
    indirect_po_pso_attainment = None
    final_po_pso_attainment = None

    if indirect_attainment:
        try:
            indirect_po_pso_attainment = ece_orignal_updated.compute_po_pso_weighted_avg(
                pd.Series(indirect_attainment), co_po_mapping_df
            )
            final_po_pso_attainment = (direct_po_pso_attainment * 0.90) + (indirect_po_pso_attainment * 0.10)
        except Exception as e:
            print(f"Error computing indirect PO/PSO attainment: {e}")

    intermediate['direct_po_pso_attainment'] = direct_po_pso_attainment.to_dict()
    intermediate['indirect_attainment_values'] = indirect_attainment
    if indirect_po_pso_attainment is not None:
        intermediate['indirect_po_pso_attainment'] = indirect_po_pso_attainment.to_dict()
    if final_po_pso_attainment is not None:
        intermediate['final_po_pso_attainment'] = final_po_pso_attainment.to_dict()

    if excel_path and os.path.exists(excel_path):
        try:
            wb = load_workbook(excel_path)
            ws = wb.active

            col_map = {}
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=col_idx).value
                if val:
                    col_map[str(val).strip()] = col_idx

            label_col_idx = col_map.get('CO Stats')

            if indirect_attainment and label_col_idx:
                for row_idx in range(1, ws.max_row + 1):
                    cell_val = ws.cell(row=row_idx, column=label_col_idx).value
                    if cell_val and 'Indirect CO avg' == str(cell_val).strip():
                        for co, val in indirect_attainment.items():
                            if co in col_map:
                                ws.cell(row=row_idx, column=col_map[co], value=round(val, 4))
                        break

            if label_col_idx:
                next_row = ws.max_row + 2

                if indirect_po_pso_attainment is not None:
                    ws.cell(row=next_row, column=label_col_idx, value='Indirect PO/PSO Attainment')
                    for po, val in indirect_po_pso_attainment.items():
                        if po in col_map:
                            ws.cell(row=next_row, column=col_map[po], value=round(float(val), 4))
                    next_row += 1

                if final_po_pso_attainment is not None:
                    ws.cell(
                        row=next_row,
                        column=label_col_idx,
                        value='Final PO/PSO Attainment (90% Direct + 10% Indirect)'
                    )
                    for po, val in final_po_pso_attainment.items():
                        if po in col_map:
                            ws.cell(row=next_row, column=col_map[po], value=round(float(val), 4))

            wb.save(excel_path)
            print(f"[excel] Updated {excel_path} with indirect attainment data")
        except Exception as e:
            print(f"[excel] Error updating Excel with indirect data: {e}")

    return {
        'course_title': course_title,
        'course_filename': course_filename or os.path.basename(course_path),
        'mapping_filename': mapping_filename or os.path.basename(mapping_path),
        'intermediate': intermediate,
        'unique_COs': intermediate.get('unique_COs', []),
        'excel_path': excel_path,
    }


def infer_bulk_failed_file(error_message, course_filename, compare_filename, mapping_filename):
    message = (error_message or '').lower()
    if 'comparison row' in message or 'cell id' in message or 'excel cell' in message or compare_filename.lower() in message:
        return compare_filename or 'Comparison file'
    if 'mapping workbook' in message or 'mapping file' in message or mapping_filename.lower() in message:
        return mapping_filename or 'Mapping file'
    return course_filename or 'Input file'


def is_bulk_row_blank(row_id, form, files):
    course_path = form.get(f'course_file_path_{row_id}', '').strip()
    course_title = form.get(f'course_title_{row_id}', '').strip()
    course_file = files.get(f'course_file_{row_id}')
    compare_file = files.get(f'compare_file_{row_id}')
    has_course_file = bool(course_file and course_file.filename)
    has_compare_file = bool(compare_file and compare_file.filename)
    return not any([course_path, course_title, has_course_file, has_compare_file])


def process_bulk_eval_row(row_id, row_number, form, files, mapping_path, mapping_filename):
    cleanup_paths = []
    course_path = form.get(f'course_file_path_{row_id}', '').strip()
    course_file = files.get(f'course_file_{row_id}')
    compare_file = files.get(f'compare_file_{row_id}')
    course_title = form.get(f'course_title_{row_id}', '').strip()
    co_cell_ref = form.get(f'co_cell_{row_id}', '').strip()
    po_cell_ref = form.get(f'po_cell_{row_id}', '').strip()
    selected_programmes = form.getlist(f'programmes_{row_id}')
    selected_branches = form.getlist(f'branches_{row_id}')

    course_filename = get_display_filename(course_file, course_path)
    compare_filename = get_display_filename(compare_file)
    scope_summary = summarize_scope_selection(selected_programmes, selected_branches)
    stage = 'validation'

    try:
        if not course_title:
            raise ValueError('Course title is required.')

        if not compare_file or not allowed_file(compare_file.filename):
            raise ValueError('A valid comparison Excel (.xlsx) file is required.')

        if not course_path or not os.path.exists(course_path):
            if not course_file or not allowed_file(course_file.filename):
                raise ValueError('A valid input Excel (.xlsx) file is required.')
            course_path = save_uploaded_file(course_file, f'bulk_eval_input_{row_id}')
            cleanup_paths.append(course_path)
            course_filename = get_display_filename(course_file, course_path)
        else:
            cleanup_paths.append(course_path)

        stage = 'comparison upload'
        compare_path = save_uploaded_file(compare_file, f'bulk_eval_compare_{row_id}')
        cleanup_paths.append(compare_path)

        stage = 'calculation'
        included_rolls = build_included_rolls(course_path, selected_programmes, selected_branches)
        evaluation = build_evaluation_payload(
            course_path,
            mapping_path,
            course_title,
            compare_path,
            included_rolls=included_rolls,
            co_cell_ref=co_cell_ref,
            po_cell_ref=po_cell_ref,
        )

        generated_excel_path = evaluation.get('intermediate', {}).get('excel_path')
        if generated_excel_path:
            cleanup_paths.append(generated_excel_path)

        return {
            'status': 'success',
            'row_number': row_number,
            'course_title': course_title,
            'course_filename': course_filename,
            'compare_filename': compare_filename,
            'mapping_filename': mapping_filename,
            'scope_summary': scope_summary,
            'co_table': evaluation['co_table'],
            'po_table': evaluation['po_table'],
        }
    except Exception as e:
        error_message = str(e)
        return {
            'status': 'error',
            'row_number': row_number,
            'course_title': course_title or '(not selected)',
            'course_filename': course_filename or '(missing)',
            'compare_filename': compare_filename or '(missing)',
            'mapping_filename': mapping_filename,
            'scope_summary': scope_summary,
            'error_stage': stage,
            'error_message': error_message,
            'failed_file': infer_bulk_failed_file(error_message, course_filename, compare_filename, mapping_filename),
        }
    finally:
        for path in cleanup_paths:
            remove_file_if_exists(path)


def parse_student_rolls(file_path):
    """Parse an uploaded course file and categorize students by programme and branch.
    Also extracts unique COs found in the CO row of the file.
    
    Attempts to read a 'Branch' column from the raw data (e.g. 'July 2022/BTech/ECE-IIITD/Semester 7')
    to determine the actual programme and branch. Falls back to roll-number prefix heuristics.
    """
    # --- Read raw (no header/index) to detect the Branch column ---
    raw = pd.read_excel(file_path, header=None)
    
    # Find the row where 'Branch' appears, and which column it's in
    branch_col_idx = None
    data_start_row = None
    for r in range(min(10, len(raw))):
        for c in range(min(5, raw.shape[1])):
            val = str(raw.iloc[r, c]).strip()
            if val.lower() == 'branch':
                branch_col_idx = c
                data_start_row = r + 1  # data starts after 'Branch' label row
                break
        if branch_col_idx is not None:
            break
    
    # Build roll -> branch_string mapping from raw data
    roll_branch_map = {}  # roll_number -> raw branch string
    if branch_col_idx is not None and data_start_row is not None:
        roll_col_idx = 0  # first column has roll numbers
        for r in range(data_start_row, len(raw)):
            roll = str(raw.iloc[r, roll_col_idx]).strip()
            branch_str = str(raw.iloc[r, branch_col_idx]).strip()
            # Unescape HTML entities
            branch_str = branch_str.replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>')
            if roll and roll.lower() not in ('nan', ''):
                roll_branch_map[roll] = branch_str
    
    # --- Now read with standard header/index for CO extraction ---
    df = pd.read_excel(file_path, header=0, index_col=0)
    df.dropna(how='all', inplace=True)
    df.index = df.index.map(lambda x: str(x).strip() if isinstance(x, str) else x)

    # Accept Max_Marks_scaled as alias for Max_Marks
    if 'Max_Marks_scaled' in df.index and 'Max_Marks' not in df.index:
        df.rename(index={'Max_Marks_scaled': 'Max_Marks'}, inplace=True)

    # Extract COs from the CO row
    cos = []
    if 'CO' in df.index:
        for val in df.loc['CO']:
            if isinstance(val, str):
                for co in val.split(','):
                    co = co.strip()
                    if co and re.match(r'^CO\d+$', co) and co not in cos:
                        cos.append(co)
    cos.sort(key=lambda x: int(re.findall(r'\d+', x)[0]) if re.findall(r'\d+', x) else 0)

    metadata_labels = {'CO', 'Max_Marks', 'Max_Marks_scaled', 'Roll No.', 'Roll No', 'Branch'}
    all_rolls = [str(idx).strip() for idx in df.index
                 if not pd.isna(idx) and str(idx).strip() not in metadata_labels]

    programmes = {}  # e.g. {'UG': {'count': 10, 'rolls': [...]}, 'PG': {...}}
    branches = {}    # e.g. {'ECE': {'count': 5, 'rolls': [...]}, ...}

    def _extract_branch_from_string(branch_str):
        """Parse branch string like 'July 2022/BTech/ECE-IIITD/Semester 7' or
        'July 2024/MTech (ECE)/VLSI & ES-IIITD/Semester 3'."""
        if not branch_str or branch_str.lower() == 'nan':
            return None, None
        parts = [p.strip() for p in branch_str.split('/')]
        prog = None
        branch = None
        for p in parts:
            pl = p.lower()
            if 'btech' in pl or 'b.tech' in pl:
                prog = 'UG'
            elif 'mtech' in pl or 'm.tech' in pl:
                prog = 'PG'
            elif 'phd' in pl:
                prog = 'PhD'
        # Extract branch: look for the part after BTech/MTech that contains branch name
        # e.g. 'ECE-IIITD' -> 'ECE', 'VLSI & ES-IIITD' -> 'VLSI & ES'
        for p in parts:
            if '-IIITD' in p:
                branch = p.replace('-IIITD', '').replace('/IIITD', '').strip()
                break
            elif p.upper().startswith('ECE') or p.upper().startswith('CSE') or p.upper().startswith('CS'):
                branch = p.strip()
                break
        return prog, branch

    for roll in all_rolls:
        roll_upper = roll.upper()
        prog = None
        branch = None
        
        # Try to get programme/branch from file's Branch column
        if roll in roll_branch_map:
            prog, branch = _extract_branch_from_string(roll_branch_map[roll])
        
        # Fall back to roll-number prefix heuristics
        if prog is None:
            if roll_upper.startswith('MT'):
                prog = 'PG'
            elif roll_upper.startswith('PHD'):
                prog = 'PhD'
            elif re.match(r'^\d{7}$', roll):
                prog = 'UG'
            else:
                prog = 'Other'
        
        programmes.setdefault(prog, {'count': 0, 'rolls': []})
        programmes[prog]['count'] += 1
        programmes[prog]['rolls'].append(roll)
        
        # Record branch keyed by (prog, branch) to avoid mixing programmes
        if branch:
            branch_key = f"{prog}::{branch}"
            branches.setdefault(branch_key, {'count': 0, 'rolls': [], 'programme': prog, 'branch': branch})
            branches[branch_key]['count'] += 1
            branches[branch_key]['rolls'].append(roll)

    # Flatten branches for the return value: use branch_key as key
    return {
        'cos': cos,
        'programmes': {k: v['count'] for k, v in programmes.items()},
        'branches': {k: {'count': v['count'], 'programme': v['programme'], 'branch': v['branch']}
                     for k, v in branches.items()},
        'total_students': len(all_rolls),
        'rolls_by_programme': {k: v['rolls'] for k, v in programmes.items()},
        'rolls_by_branch': {k: v['rolls'] for k, v in branches.items()},
    }


def extract_course_names(mapping_path):
    """Read a CO-PO mapping Excel and return the list of course names found."""
    sheet_preference = [
        "Course outcome mapping UG",
        "CO mapping - PG",
        "Course mapping UG",
        "Course mapping PG",
    ]
    courses = []
    try:
        xl = pd.ExcelFile(mapping_path)
        for sheet in sheet_preference:
            if sheet not in xl.sheet_names:
                continue
            df = pd.read_excel(mapping_path, sheet_name=sheet)
            if df.shape[1] < 16:
                continue
            first_col = df.columns[0]
            for val in df[first_col].astype(str):
                val = val.strip()
                if not val or val.lower() == "nan":
                    continue
                # Skip CO rows (CO1, CO2, ...) and the header row
                if re.match(r'^CO\d+', val):
                    continue
                # Skip PO/PSO header-like rows
                if re.match(r'^(PO|PSO)\d*$', val):
                    continue
                # Keep anything that looks like a course name
                if len(val) > 3:
                    courses.append(val)
            if courses:
                break
    except Exception as e:
        print(f"[extract_course_names] Error reading {mapping_path}: {e}")
    return courses


@app.route('/api/course_names', methods=['GET', 'POST'])
def api_course_names():
    """Return course names from the mapping file as JSON.
    GET  -> reads from the default mapping file
    POST -> reads from an uploaded custom mapping file
    """
    if request.method == 'POST':
        f = request.files.get('mapping_file')
        if not f or not allowed_file(f.filename):
            return jsonify({'error': 'Invalid file'}), 400
        path = os.path.join(app.config['UPLOAD_FOLDER'], 'tmp_mapping_' + f.filename)
        f.save(path)
        try:
            names = extract_course_names(path)
        finally:
            if os.path.exists(path):
                os.remove(path)
        return jsonify({'courses': names})

    # GET — use default mapping
    default_path = MAPPING_FILES['CO-PO Mapping Nov 25']
    names = extract_course_names(default_path)
    return jsonify({'courses': names})


@app.route('/api/clear_uploads', methods=['POST'])
def api_clear_uploads():
    """Delete uploaded files and clear in-memory result/download contexts."""
    removed = 0
    for filename in os.listdir(UPLOAD_FOLDER):
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        if os.path.isfile(filepath):
            try:
                os.remove(filepath)
                removed += 1
            except Exception as e:
                print(f"[clear_uploads] Could not remove {filepath}: {e}")
    cleared_contexts = clear_runtime_contexts()
    return jsonify({
        'removed': removed,
        'cleared_result_contexts': cleared_contexts['result_contexts'],
        'cleared_download_contexts': cleared_contexts['download_contexts'],
    })


@app.route('/api/parse_students', methods=['POST'])
def api_parse_students():
    """Upload and parse a course file to detect programmes, branches, and COs."""
    f = request.files.get('course_file')
    if not f or not allowed_file(f.filename):
        return jsonify({'error': 'Invalid file'}), 400

    saved_path = save_uploaded_file(f, 'parsed_input')

    try:
        result = parse_student_rolls(saved_path)
    except Exception as e:
        if os.path.exists(saved_path):
            os.remove(saved_path)
        return jsonify({'error': str(e)}), 500

    return jsonify({
        'cos': result['cos'],
        'programmes': result['programmes'],
        'branches': result['branches'],
        'total_students': result['total_students'],
        'file_path': saved_path,
    })


def extract_cos_for_course(mapping_path, course_title):
    """Extract the list of CO labels for a given course from the CO-PO mapping file."""
    sheet_preference = [
        "Course outcome mapping UG",
        "CO mapping - PG",
        "Course mapping UG",
        "Course mapping PG",
    ]
    try:
        xl = pd.ExcelFile(mapping_path)
        for sheet in sheet_preference:
            if sheet not in xl.sheet_names:
                continue
            df = pd.read_excel(mapping_path, sheet_name=sheet)
            if df.shape[1] < 16:
                continue
            first_col = df.columns[0]
            col0 = df[first_col].astype(str)
            hits = [i for i, v in enumerate(col0.values) if course_title.lower() in v.lower()]
            if not hits:
                continue
            start = hits[0]
            co_labels = []
            i = start + 1
            while i < len(df):
                v = str(df.loc[i, first_col]).strip()
                if v.upper().startswith("CO") and re.match(r'^CO\d+', v, re.IGNORECASE):
                    co_labels.append(v)
                    i += 1
                    continue
                if v.lower() == "nan" or v == "" or re.match(r"^\s*ECE-\d+", v):
                    break
                break
            if co_labels:
                return co_labels
    except Exception as e:
        print(f"[extract_cos_for_course] Error: {e}")
    return []


def lookup_indirect_values(indirect_path, course_title, co_labels):
    """Look up existing indirect CO attainment values from the indirect file."""
    values = {}
    if not os.path.exists(indirect_path):
        return values
    try:
        indirect_df = pd.read_excel(indirect_path)
        match_row = None
        for idx, val in indirect_df.iloc[:, 0].items():
            if isinstance(val, str) and course_title.lower() in val.lower():
                match_row = idx
                break
        if match_row is not None:
            for co in co_labels:
                col_name = None
                if co in indirect_df.columns:
                    col_name = co
                else:
                    m = re.search(r'(\d+)', co)
                    if m:
                        co_num = m.group(1)
                        alt1 = f"C{co_num.zfill(2)}"
                        alt2 = f"C{int(co_num)}"
                        if alt1 in indirect_df.columns:
                            col_name = alt1
                        elif alt2 in indirect_df.columns:
                            col_name = alt2
                if col_name:
                    val = indirect_df.loc[match_row, col_name]
                    if pd.notna(val):
                        values[co] = float(val)
    except Exception as e:
        print(f"[lookup_indirect_values] Error: {e}")
    return values


@app.route('/api/course_cos', methods=['POST'])
def api_course_cos():
    """Return the list of COs for a course and any existing indirect attainment values."""
    course_title = request.form.get('course_title', '').strip()
    if not course_title:
        return jsonify({'error': 'No course title provided'}), 400

    # Determine mapping path
    mapping_file = request.files.get('mapping_file')
    mapping_option = request.form.get('mapping_option', 'default')
    tmp_mapping_path = None
    if mapping_option == 'upload' and mapping_file and allowed_file(mapping_file.filename):
        tmp_mapping_path = os.path.join(app.config['UPLOAD_FOLDER'], 'tmp_cos_mapping_' + mapping_file.filename)
        mapping_file.save(tmp_mapping_path)
        mapping_path = tmp_mapping_path
    else:
        mapping_path = MAPPING_FILES['CO-PO Mapping Nov 25']

    try:
        co_labels = extract_cos_for_course(mapping_path, course_title)
        indirect_values = lookup_indirect_values(DEFAULT_INDIRECT_PATH, course_title, co_labels)
    finally:
        if tmp_mapping_path and os.path.exists(tmp_mapping_path):
            os.remove(tmp_mapping_path)

    return jsonify({
        'cos': co_labels,
        'indirect_values': indirect_values,
        'found_in_file': bool(indirect_values)
    })


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        course_title = request.form.get('course_title')
        course_file = request.files.get('course_file')
        
        # Get course file path (pre-uploaded via /api/parse_students)
        course_path = request.form.get('course_file_path', '').strip()
        
        # If no pre-uploaded path, check for file upload
        if not course_path or not os.path.exists(course_path):
            if not course_file or not allowed_file(course_file.filename):
                flash('Please upload a valid course data Excel (.xlsx) file.')
                return redirect(request.url)
            course_path = os.path.join(app.config['UPLOAD_FOLDER'], course_file.filename)
            course_file.save(course_path)
        
        if not course_title:
            flash('Please enter the course title.')
            return redirect(request.url)
        
        # Handle CO-PO mapping file
        mapping_option = request.form.get('mapping_option')
        custom_mapping = None
        if mapping_option == 'upload':
            custom_mapping = request.files.get('custom_mapping_file')
            if not custom_mapping or not allowed_file(custom_mapping.filename):
                flash('Please upload a valid CO-PO mapping Excel (.xlsx) file.')
                return redirect(request.url)
            mapping_path = os.path.join(app.config['UPLOAD_FOLDER'], 'custom_mapping_' + custom_mapping.filename)
            custom_mapping.save(mapping_path)
        else:
            mapping_path = MAPPING_FILES['CO-PO Mapping Nov 25']
        
        # --- Build included_rolls from programme/branch selections ---
        selected_programmes = request.form.getlist('programmes')
        selected_branches = request.form.getlist('branches')
        
        included_rolls = build_included_rolls(course_path, selected_programmes, selected_branches)
        
        # Collect indirect CO attainment values from text inputs
        indirect_attainment_input = {}
        for key, val in request.form.items():
            if key.startswith('indirect_co_'):
                co_name = key.replace('indirect_co_', '')
                val = val.strip()
                if val:
                    try:
                        indirect_attainment_input[co_name] = float(val)
                    except ValueError:
                        flash(f'Invalid value for {co_name}: "{val}". Please enter a number.')
                        return redirect(request.url)
        
        target_value = 50

        cleanup_input_paths = []
        if course_path and os.path.abspath(course_path).startswith(os.path.abspath(UPLOAD_FOLDER)):
            cleanup_input_paths.append(course_path)
        if mapping_path and os.path.abspath(mapping_path).startswith(os.path.abspath(UPLOAD_FOLDER)):
            cleanup_input_paths.append(mapping_path)

        try:
            payload = prepare_results_payload(
                course_path,
                mapping_path,
                course_title,
                included_rolls=included_rolls,
                indirect_attainment=indirect_attainment_input,
                course_filename=get_display_filename(course_file, course_path),
                mapping_filename=get_display_filename(custom_mapping, mapping_path),
                target_value=target_value,
            )
            payload['download_id'] = store_download_context(payload.get('excel_path'))
            result_id = store_result_context(payload)
        except ValueError as e:
            flash(str(e))
            return redirect(request.url)
        except Exception as e:
            flash(f'Error processing file: {e}')
            return redirect(request.url)
        finally:
            for path in cleanup_input_paths:
                remove_file_if_exists(path)

        return redirect(url_for('results', result_id=result_id))
    return render_template('index.html')


@app.route('/eval', methods=['GET', 'POST'])
def eval_space():
    if request.method == 'POST':
        course_title = request.form.get('course_title', '').strip()
        if not course_title:
            flash('Please select a course title.')
            return redirect(url_for('eval_space'))

        course_file = request.files.get('course_file')
        compare_file = request.files.get('compare_file')
        if not compare_file or not allowed_file(compare_file.filename):
            flash('Please upload a valid "to compare with" Excel (.xlsx) file.')
            return redirect(url_for('eval_space'))

        cleanup_paths = []
        evaluation = None
        error_message = None

        course_path = request.form.get('course_file_path', '').strip()
        if not course_path or not os.path.exists(course_path):
            if not course_file or not allowed_file(course_file.filename):
                flash('Please upload a valid input Excel (.xlsx) file.')
                return redirect(url_for('eval_space'))
            course_path = save_uploaded_file(course_file, 'eval_input')
            cleanup_paths.append(course_path)
            course_filename = course_file.filename
        else:
            cleanup_paths.append(course_path)
            course_filename = course_file.filename if course_file and course_file.filename else os.path.basename(course_path)

        compare_filename = compare_file.filename
        mapping_option = request.form.get('mapping_option')
        mapping_filename = os.path.basename(MAPPING_FILES['CO-PO Mapping Nov 25'])
        selected_programmes = request.form.getlist('programmes')
        selected_branches = request.form.getlist('branches')
        co_cell_ref = request.form.get('co_attainment_cell', '').strip()
        po_cell_ref = request.form.get('po_attainment_cell', '').strip()
        included_rolls = build_included_rolls(course_path, selected_programmes, selected_branches)
        scope_summary = summarize_scope_selection(selected_programmes, selected_branches)

        try:
            compare_path = save_uploaded_file(compare_file, 'eval_compare')
            cleanup_paths.append(compare_path)

            if mapping_option == 'upload':
                custom_mapping = request.files.get('custom_mapping_file')
                if not custom_mapping or not allowed_file(custom_mapping.filename):
                    raise ValueError('Please upload a valid CO-PO mapping Excel (.xlsx) file.')
                mapping_path = save_uploaded_file(custom_mapping, 'eval_mapping')
                cleanup_paths.append(mapping_path)
                mapping_filename = custom_mapping.filename
            else:
                mapping_path = MAPPING_FILES['CO-PO Mapping Nov 25']

            evaluation = build_evaluation_payload(
                course_path,
                mapping_path,
                course_title,
                compare_path,
                included_rolls=included_rolls,
                co_cell_ref=co_cell_ref,
                po_cell_ref=po_cell_ref,
            )
            generated_excel_path = evaluation.get('intermediate', {}).get('excel_path')
            if generated_excel_path:
                cleanup_paths.append(generated_excel_path)

        except ValueError as e:
            error_message = str(e)
        except Exception as e:
            error_message = f'Error comparing files: {e}'

        for path in cleanup_paths:
            remove_file_if_exists(path)

        if error_message:
            flash(error_message)
            return redirect(url_for('eval_space'))

        return render_template(
            'eval_results.html',
            course_title=course_title,
            course_filename=course_filename,
            compare_filename=compare_filename,
            mapping_filename=mapping_filename,
            target_value=50,
            scope_summary=scope_summary,
            co_table=evaluation['co_table'],
            po_table=evaluation['po_table'],
        )

    return render_template('eval.html')


@app.route('/eval/bulk', methods=['GET', 'POST'])
def bulk_eval_space():
    if request.method == 'POST':
        row_ids = request.form.getlist('row_ids')
        if not row_ids:
            flash('Please add at least one comparison row.')
            return redirect(url_for('bulk_eval_space'))

        mapping_option = request.form.get('mapping_option')
        mapping_filename = os.path.basename(MAPPING_FILES['CO-PO Mapping Nov 25'])
        mapping_cleanup_path = None

        try:
            if mapping_option == 'upload':
                custom_mapping = request.files.get('bulk_custom_mapping_file')
                if not custom_mapping or not allowed_file(custom_mapping.filename):
                    flash('Please upload a valid CO-PO mapping Excel (.xlsx) file for bulk evaluation.')
                    return redirect(url_for('bulk_eval_space'))
                mapping_cleanup_path = save_uploaded_file(custom_mapping, 'bulk_eval_mapping')
                mapping_path = mapping_cleanup_path
                mapping_filename = custom_mapping.filename
            else:
                mapping_path = MAPPING_FILES['CO-PO Mapping Nov 25']

            results_data = []
            row_number = 0
            for row_id in row_ids:
                if is_bulk_row_blank(row_id, request.form, request.files):
                    continue
                row_number += 1
                results_data.append(
                    process_bulk_eval_row(
                        row_id,
                        row_number,
                        request.form,
                        request.files,
                        mapping_path,
                        mapping_filename,
                    )
                )

            if not results_data:
                flash('Please fill at least one bulk evaluation row before submitting.')
                return redirect(url_for('bulk_eval_space'))

            success_count = sum(1 for row in results_data if row['status'] == 'success')
            error_count = sum(1 for row in results_data if row['status'] == 'error')

            return render_template(
                'eval_bulk_results.html',
                results_data=results_data,
                mapping_filename=mapping_filename,
                total_rows=len(results_data),
                success_count=success_count,
                error_count=error_count,
            )
        finally:
            remove_file_if_exists(mapping_cleanup_path)

    return render_template('eval_bulk.html')


@app.route('/results/<result_id>', methods=['GET'])
def results(result_id):
    payload = get_result_context(result_id)
    if not payload:
        flash('Results not found or expired. Please generate them again.')
        return redirect(url_for('index'))

    return render_template(
        'results.html',
        course_title=payload['course_title'],
        course_filename=payload['course_filename'],
        mapping_filename=payload['mapping_filename'],
        intermediate=payload['intermediate'],
        unique_COs=payload['unique_COs'],
        download_id=payload.get('download_id'),
    )

if __name__ == '__main__':
    app.run(debug=True, port=8000)
