from __future__ import annotations

import io
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.course_allocation.services.allocation_faculty_resolver import (
    is_placeholder_name,
    refresh_alias_cache,
    resolve_allocation_faculty,
)
from app.course_allocation.services.semester_service import academic_year_for_semester
from app.utils.contribution_faculty_resolver import FacultyResolveResult

_SEMESTER_BANNER_RE = re.compile(r"^(Monsoon|Winter)\s+(\d{4})\s+Semester", re.I)
_HEADER_MARKERS = ("course code", "course name", "faculty")


def _derive_ug_pg_core(label: str) -> tuple[str, str]:
    text = (label or "").strip()
    upper = text.upper()
    ug = "UG" in upper or "B.TECH" in upper or "BTECH" in upper
    pg = "PG" in upper or "M.TECH" in upper or "MTECH" in upper or "M. TECH" in upper
    if ug and pg:
        ug_pg = "UG/PG"
    elif pg:
        ug_pg = "PG"
    elif ug:
        ug_pg = "UG"
    else:
        ug_pg = "UG/PG"
    core = "CORE" in upper
    elective = "ELECTIVE" in upper
    if core and elective:
        ce = "Core/Elective"
    elif core:
        ce = "Core"
    elif elective:
        ce = "Elective"
    else:
        ce = "Elective"
    return ug_pg, ce


_FIRST_YEAR_KEYWORDS = [
    ("introduction to programming", "Introduction to Programming"),
    ("data structures", "Data Structures and Algorithms"),
    ("digital circuits", "Digital Circuits"),
    ("basic electronics", "Basic Electronics"),
    ("linear algebra", "Maths I (Linear Algebra)"),
    ("probability", "Maths II (Probability & Statistics)"),
    ("hci", "Introduction to HCI"),
    ("computer organisation", "Computer Organization"),
    ("computer organization", "Computer Organization"),
    ("communication skills", "Communication Skills"),
]


def _detect_first_year(course_name: str) -> tuple[bool, str | None]:
    lower = course_name.lower()
    if "internet of things" in lower or ("iot" in lower and "introduction to programming" in lower):
        return False, None
    for keyword, canonical in _FIRST_YEAR_KEYWORDS:
        if keyword in lower:
            return True, canonical
    return False, None


def parse_allocation_xlsx(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_out: list[dict] = []
    current_semester: str | None = None
    header_map: dict[str, int] | None = None

    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not any(cells):
            continue
        joined = " ".join(cells).strip()
        banner = _SEMESTER_BANNER_RE.match(joined)
        if banner:
            current_semester = f"{banner.group(1).title()} {banner.group(2)}"
            header_map = None
            continue
        lower_joined = joined.lower()
        if "course code" in lower_joined and "faculty" in lower_joined:
            header_map = {}
            for idx, cell in enumerate(cells):
                cl = cell.lower()
                if "course code" in cl:
                    header_map["code"] = idx
                elif "course name" in cl:
                    header_map["name"] = idx
                elif "faculty" in cl:
                    header_map["faculty"] = idx
                elif "core" in cl or "elective" in cl:
                    header_map["label"] = idx
            continue
        if not current_semester or not header_map:
            continue
        code = cells[header_map.get("code", 0)] if header_map.get("code") is not None else ""
        name = cells[header_map.get("name", 1)] if header_map.get("name") is not None else ""
        from app.course_allocation.services.course_identity_resolver import split_course_code_name

        code, name = split_course_code_name(code, name)
        faculty = cells[header_map.get("faculty", 2)] if header_map.get("faculty") is not None else ""
        label = cells[header_map.get("label", 3)] if header_map.get("label") is not None else ""
        if not code or code.lower() == "course code":
            continue
        ug_pg, core_elective = _derive_ug_pg_core(label)
        is_fy, fy_name = _detect_first_year(name)
        placeholder = is_placeholder_name(faculty)
        rows_out.append(
            {
                "faculty_name": faculty,
                "semester": current_semester,
                "academic_year": academic_year_for_semester(current_semester),
                "course_code": code,
                "course_name": name,
                "ug_pg": ug_pg,
                "core_elective": core_elective,
                "is_first_year": is_fy,
                "first_year_course_name": fy_name,
                "source": "new",
                "is_faculty_placeholder": placeholder,
            }
        )
    wb.close()
    return rows_out


def preview_upload(path: Path, db: Session | None = None) -> dict:
    parsed = parse_allocation_xlsx(path)
    semesters = sorted({r["semester"] for r in parsed})

    real_names = sorted(
        {
            r["faculty_name"].strip()
            for r in parsed
            if not r["is_faculty_placeholder"] and (r.get("faculty_name") or "").strip()
        }
    )

    # Resolve each distinct name against the live faculty table / alias map so the
    # preview reflects what will actually happen on commit. Without a db session we
    # cannot match, so report nothing as unmatched rather than flag every name.
    matched_names: list[str] = []
    unmatched_names: list[str] = []
    if db is not None:
        refresh_alias_cache(db)
        for name in real_names:
            resolved = resolve_allocation_faculty(db, name)
            if isinstance(resolved, FacultyResolveResult):
                matched_names.append(name)
            else:
                unmatched_names.append(name)

    return {
        "row_count": len(parsed),
        "semesters": semesters,
        "rows": parsed[:50],
        "truncated": len(parsed) > 50,
        "matched_count": len(matched_names),
        "unmatched_names": unmatched_names,
        "errors": [],
    }
