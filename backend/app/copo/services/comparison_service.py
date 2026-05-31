"""Excel comparison helpers for validation workflow (legacy parity)."""

import os
import re

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string

from app.copo.constants import PO_PSO_HEADERS


def normalize_summary_label(value) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def is_co_summary_label(value: str) -> bool:
    normalized = normalize_summary_label(value)
    return normalized.startswith("max50mean05std") and "weightedavg" not in normalized


def is_po_summary_label(value: str) -> bool:
    normalized = normalize_summary_label(value)
    return normalized.startswith("weightedavgusingco") and "mean05std" in normalized


def coerce_numeric(value):
    if value is None:
        return None
    if isinstance(value, (int, float, np.number)):
        if pd.isna(value):
            return None
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def find_summary_anchor(worksheet, matcher):
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and matcher(cell.value):
                return cell
    return None


def find_summary_anchor_in_workbook(workbook, matcher):
    for worksheet in workbook.worksheets:
        anchor_cell = find_summary_anchor(worksheet, matcher)
        if anchor_cell is not None:
            return worksheet, anchor_cell
    return None, None


def normalize_cell_reference(cell_ref: str | None) -> str:
    if cell_ref is None:
        return ""
    normalized = str(cell_ref).strip().upper()
    if not normalized:
        return ""
    coordinate_from_string(normalized)
    return normalized


def find_numeric_start_column(worksheet, row_idx, start_col):
    for col_idx in range(start_col, worksheet.max_column + 1):
        if coerce_numeric(worksheet.cell(row_idx, col_idx).value) is not None:
            return col_idx
    return None


def extract_consecutive_numeric_values(worksheet, row_idx, start_col, count, context_label):
    extracted = []
    missing_cells = []
    for offset in range(count):
        col_idx = start_col + offset
        cell = worksheet.cell(row_idx, col_idx)
        cell_value = cell.value
        numeric_value = coerce_numeric(cell_value)
        if cell_value in (None, "") or numeric_value is None:
            missing_cells.append(cell.coordinate)
        extracted.append(numeric_value)
    if missing_cells:
        raise ValueError(
            f"Expected {count} values from {context_label}, but these cells were blank or non-numeric: "
            + ", ".join(missing_cells)
        )
    return extracted


def read_values_from_anchor_cell(worksheet, anchor_cell, count, context_label):
    if coerce_numeric(anchor_cell.value) is not None:
        start_col = anchor_cell.column
    else:
        start_col = find_numeric_start_column(worksheet, anchor_cell.row, anchor_cell.column + 1)
    if start_col is None:
        raise ValueError(
            f"Could not find a numeric starting point on row {anchor_cell.row} for {context_label}."
        )
    return extract_consecutive_numeric_values(
        worksheet, anchor_cell.row, start_col, count, context_label
    )


def read_summary_values_from_output(
    output_path: str,
    expected_headers: list[str],
    matcher,
    cell_ref: str | None = None,
    fixed_count: int | None = None,
):
    workbook = load_workbook(output_path, data_only=True)
    value_count = fixed_count if fixed_count is not None else len(expected_headers)
    override_ref = normalize_cell_reference(cell_ref) if cell_ref else ""
    if override_ref:
        errors = []
        for worksheet in workbook.worksheets:
            anchor_cell = worksheet[override_ref]
            matched_label = f"{worksheet.title}!{override_ref}"
            try:
                sequential_values = read_values_from_anchor_cell(
                    worksheet, anchor_cell, value_count, matched_label
                )
                return dict(zip(expected_headers, sequential_values)), matched_label
            except ValueError as e:
                errors.append(f"{worksheet.title}: {e}")
        raise ValueError(
            "Could not read the override cell across workbook sheets. " + " | ".join(errors)
        )
    worksheet, anchor_cell = find_summary_anchor_in_workbook(workbook, matcher)
    if anchor_cell is None:
        raise ValueError(
            f'Could not locate the required comparison row in "{os.path.basename(output_path)}".'
        )
    matched_label = f"{worksheet.title}: {str(anchor_cell.value).strip()}"
    sequential_values = read_values_from_anchor_cell(
        worksheet, anchor_cell, value_count, matched_label
    )
    return dict(zip(expected_headers, sequential_values)), matched_label


def build_delta_values(calculated_values, read_values, ordered_headers):
    delta = {}
    for header in ordered_headers:
        calc_val = calculated_values.get(header)
        read_val = read_values.get(header)
        if calc_val is None or read_val is None:
            delta[header] = None
        else:
            delta[header] = float(calc_val) - float(read_val)
    return delta


def build_eval_table(title, columns, calculated_values, read_values, matched_label):
    delta_values = build_delta_values(calculated_values, read_values, columns)
    return {
        "title": title,
        "matched_label": matched_label,
        "columns": columns,
        "rows": [
            {
                "label": "Calculated",
                "kind": "calculated",
                "values": [calculated_values.get(col) for col in columns],
            },
            {
                "label": "Read From Output",
                "kind": "read",
                "values": [read_values.get(col) for col in columns],
            },
            {
                "label": "Delta (Calculated - Read)",
                "kind": "delta",
                "values": [delta_values.get(col) for col in columns],
            },
        ],
    }


def infer_bulk_failed_file(
    error_message: str, course_filename: str, compare_filename: str, mapping_filename: str
) -> str:
    message = (error_message or "").lower()
    if (
        "comparison row" in message
        or "cell id" in message
        or "excel cell" in message
        or compare_filename.lower() in message
    ):
        return compare_filename or "Comparison file"
    if (
        "mapping workbook" in message
        or "mapping file" in message
        or mapping_filename.lower() in message
    ):
        return mapping_filename or "Mapping file"
    return course_filename or "Input file"
