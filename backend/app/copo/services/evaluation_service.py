"""CO-PO evaluation orchestration — wraps legacy engine without rewriting formulas."""

import os

import pandas as pd
from openpyxl import load_workbook

from app.copo.constants import EVAL_CO_TABLE_TITLE, EVAL_PO_TABLE_TITLE, PO_PSO_HEADERS
from app.copo.engine import compute_po_pso_weighted_avg, main_process
from app.copo.services.comparison_service import (
    build_eval_table,
    coerce_numeric,
    is_co_summary_label,
    is_po_summary_label,
    read_summary_values_from_output,
)
from app.copo.services.file_manager import finalize_result_workbook, result_filename
from app.copo.services.marks_normalizer import cleanup_normalized_workbook, resolve_marks_workbook
from app.copo.services.student_parser import build_included_rolls, summarize_scope_selection


def prepare_results_payload(
    course_path: str,
    mapping_path: str,
    course_title: str,
    included_rolls: list[str] | None = None,
    indirect_attainment: dict[str, float] | None = None,
    course_filename: str | None = None,
    mapping_filename: str | None = None,
    target_value: int = 50,
) -> dict:
    indirect_attainment = _normalize_indirect_attainment(indirect_attainment or {})
    original_course_path = course_path
    course_path = resolve_marks_workbook(course_path)
    try:
        intermediate = main_process(
            course_path,
            mapping_path,
            course_title,
            included_rolls=included_rolls,
            target_value=target_value,
        )
    finally:
        cleanup_normalized_workbook(course_path, original_course_path)

    excel_path = intermediate.get("excel_path", "")
    co_po_mapping_df = pd.DataFrame.from_dict(intermediate["CO_PO_mapping"], orient="index")

    direct_po_pso_attainment = pd.Series(intermediate["po_pso_attainment"])
    indirect_po_pso_attainment = None
    final_po_pso_attainment = None

    if indirect_attainment:
        try:
            indirect_po_pso_attainment = compute_po_pso_weighted_avg(
                pd.Series(indirect_attainment), co_po_mapping_df
            )
            final_po_pso_attainment = (direct_po_pso_attainment * 0.90) + (
                indirect_po_pso_attainment * 0.10
            )
        except Exception as exc:
            print(f"Error computing indirect PO/PSO attainment: {exc}")

    intermediate["direct_po_pso_attainment"] = direct_po_pso_attainment.to_dict()
    intermediate["indirect_attainment_values"] = indirect_attainment
    if indirect_po_pso_attainment is not None:
        intermediate["indirect_po_pso_attainment"] = indirect_po_pso_attainment.to_dict()
    if final_po_pso_attainment is not None:
        intermediate["final_po_pso_attainment"] = final_po_pso_attainment.to_dict()

    if excel_path and os.path.exists(excel_path):
        _write_indirect_rows_to_excel(excel_path, indirect_attainment, indirect_po_pso_attainment, final_po_pso_attainment)
        excel_path = finalize_result_workbook(excel_path, course_title)

    return {
        "course_title": course_title,
        "course_filename": course_filename or os.path.basename(course_path),
        "mapping_filename": mapping_filename or os.path.basename(mapping_path),
        "intermediate": _serialize_intermediate(intermediate),
        "unique_COs": intermediate.get("unique_COs", []),
        "co_warnings": intermediate.get("co_warnings", []),
        "excel_path": excel_path,
        "download_filename": result_filename(course_title),
    }


def _write_indirect_rows_to_excel(excel_path, indirect_attainment, indirect_po, final_po):
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        col_map = {}
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col_idx).value
            if val:
                col_map[str(val).strip()] = col_idx
        label_col_idx = col_map.get("CO Stats")
        if indirect_attainment and label_col_idx:
            for row_idx in range(1, ws.max_row + 1):
                cell_val = ws.cell(row=row_idx, column=label_col_idx).value
                if cell_val and "Indirect CO avg" == str(cell_val).strip():
                    for co, val in indirect_attainment.items():
                        if co in col_map:
                            ws.cell(row=row_idx, column=col_map[co], value=round(val, 4))
                    break
        if label_col_idx:
            next_row = ws.max_row + 2
            if indirect_po is not None:
                ws.cell(row=next_row, column=label_col_idx, value="Indirect PO/PSO Attainment")
                for po, val in indirect_po.items():
                    if po in col_map:
                        ws.cell(row=next_row, column=col_map[po], value=round(float(val), 4))
                next_row += 1
            if final_po is not None:
                ws.cell(
                    row=next_row,
                    column=label_col_idx,
                    value="Final PO/PSO Attainment (90% Direct + 10% Indirect)",
                )
                for po, val in final_po.items():
                    if po in col_map:
                        ws.cell(row=next_row, column=col_map[po], value=round(float(val), 4))
        wb.save(excel_path)
    except Exception as exc:
        print(f"[excel] Error updating Excel with indirect data: {exc}")


def _normalize_indirect_attainment(indirect_attainment: dict[str, float]) -> dict[str, float]:
    """
    Normalize indirect CO attainment to percentage scale (0..100).

    Backward compatibility:
    - 0..1 inputs are interpreted as ratios and scaled to percentages.
    - 0..100 inputs are treated as already-on-percentage scale.
    """
    normalized: dict[str, float] = {}
    for co, raw_val in indirect_attainment.items():
        try:
            val = float(raw_val)
        except (TypeError, ValueError) as exc:
            raise ValueError(
                f"Invalid indirect attainment value for {co}: {raw_val!r}. Expected a number."
            ) from exc
        if val < 0:
            raise ValueError(
                f"Invalid indirect attainment value for {co}: {val}. Expected range 0 to 100."
            )
        if val <= 1:
            val *= 100.0
        if val > 100:
            raise ValueError(
                f"Invalid indirect attainment value for {co}: {val}. Expected range 0 to 100."
            )
        normalized[co] = val
    return normalized


def _serialize_intermediate(intermediate: dict) -> dict:
    """Make intermediate JSON-safe for DB storage."""
    out = dict(intermediate)
    for key in list(out.keys()):
        val = out[key]
        if hasattr(val, "to_dict"):
            out[key] = val.to_dict()
    return out


def build_evaluation_payload(
    course_path: str,
    mapping_path: str,
    course_title: str,
    compare_path: str,
    included_rolls: list[str] | None = None,
    co_cell_ref: str | None = None,
    po_cell_ref: str | None = None,
    target_value: int = 50,
) -> dict:
    original_course_path = course_path
    course_path = resolve_marks_workbook(course_path)
    try:
        intermediate = main_process(
            course_path,
            mapping_path,
            course_title,
            included_rolls=included_rolls,
            target_value=target_value,
        )
    finally:
        cleanup_normalized_workbook(course_path, original_course_path)

    co_columns = list(intermediate.get("unique_COs", []))
    calculated_co = {
        co: coerce_numeric(intermediate["CO_stats"]["pct_above"].get(co)) for co in co_columns
    }

    available_po_values = intermediate.get("po_pso_attainment", {})
    po_columns = list(PO_PSO_HEADERS)
    calculated_po = {po: coerce_numeric(available_po_values.get(po)) for po in po_columns}

    read_co, matched_co_label = read_summary_values_from_output(
        compare_path,
        co_columns,
        is_co_summary_label,
        cell_ref=co_cell_ref,
        fixed_count=len(co_columns),
    )
    read_po, matched_po_label = read_summary_values_from_output(
        compare_path,
        po_columns,
        is_po_summary_label,
        cell_ref=po_cell_ref,
        fixed_count=15,
    )

    return {
        "intermediate": _serialize_intermediate(intermediate),
        "co_warnings": intermediate.get("co_warnings", []),
        "co_table": build_eval_table(
            EVAL_CO_TABLE_TITLE, co_columns, calculated_co, read_co, matched_co_label
        ),
        "po_table": build_eval_table(
            EVAL_PO_TABLE_TITLE, po_columns, calculated_po, read_po, matched_po_label
        ),
    }


def process_bulk_row(
    course_path: str,
    compare_path: str,
    course_title: str,
    mapping_path: str,
    mapping_filename: str,
    selected_programmes: list[str],
    selected_branches: list[str],
    co_cell_ref: str = "",
    po_cell_ref: str = "",
    row_number: int = 1,
    course_filename: str = "",
    compare_filename: str = "",
) -> dict:
    scope_summary = summarize_scope_selection(selected_programmes, selected_branches)
    stage = "calculation"
    try:
        if not course_title:
            raise ValueError("Course title is required.")
        included_rolls = build_included_rolls(course_path, selected_programmes, selected_branches)
        evaluation = build_evaluation_payload(
            course_path,
            mapping_path,
            course_title,
            compare_path,
            included_rolls=included_rolls,
            co_cell_ref=co_cell_ref or None,
            po_cell_ref=po_cell_ref or None,
        )
        excel_path = evaluation.get("intermediate", {}).get("excel_path")
        return {
            "status": "success",
            "row_number": row_number,
            "course_title": course_title,
            "course_filename": course_filename,
            "compare_filename": compare_filename,
            "mapping_filename": mapping_filename,
            "scope_summary": scope_summary,
            "co_warnings": evaluation.get("co_warnings", []),
            "co_table": evaluation["co_table"],
            "po_table": evaluation["po_table"],
            "excel_path": excel_path,
        }
    except Exception as exc:
        from app.copo.services.comparison_service import infer_bulk_failed_file

        error_message = str(exc)
        return {
            "status": "error",
            "row_number": row_number,
            "course_title": course_title or "(not selected)",
            "course_filename": course_filename or "(missing)",
            "compare_filename": compare_filename or "(missing)",
            "mapping_filename": mapping_filename,
            "scope_summary": scope_summary,
            "error_stage": stage,
            "error_message": error_message,
            "failed_file": infer_bulk_failed_file(
                error_message, course_filename, compare_filename, mapping_filename
            ),
        }
