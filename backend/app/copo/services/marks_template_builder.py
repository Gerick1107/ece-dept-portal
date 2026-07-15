"""Build faculty marks Excel matching the portal sample-layout parser (Branch + merged headers)."""
from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.copo.schemas import MarksComponentSpec, MarksTemplateGenerateRequest

PRESET_COMPONENT_NAMES = [
    "Quiz",
    "MidSem",
    "EndSem",
    "Project",
    "Assignment",
    "Lab",
    "Tutorial",
    "Seminar",
    "Viva",
]

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")
_CO_FILL = PatternFill(fill_type="solid", fgColor="FFF9C4")


@dataclass(frozen=True)
class _ColumnSpec:
    group_name: str
    subheader: str
    co_cell: str | None  # None → Excel blank / parser treats as total or bonus
    is_total: bool = False
    is_bonus_question: bool = False


def _resolve_component_names(components: list[MarksComponentSpec]) -> list[tuple[str, MarksComponentSpec]]:
    """Quiz ×3 → Quiz1, Quiz2, Quiz3; single MidSem → MidSem; bonus → Name_Bonus."""
    name_counts = Counter(spec.name.strip() for spec in components)
    seen: dict[str, int] = defaultdict(int)
    resolved: list[tuple[str, MarksComponentSpec]] = []
    for spec in components:
        base = spec.name.strip()
        seen[base] += 1
        if name_counts[base] == 1:
            label = base
        else:
            label = f"{base}{seen[base]}"
        if spec.is_bonus_component:
            label = f"{label}_Bonus"
        resolved.append((label, spec))
    return resolved


def _columns_for_component(group_name: str, spec: MarksComponentSpec) -> list[_ColumnSpec]:
    if spec.questions <= 0:
        co = None if spec.is_bonus_component else ""
        return [_ColumnSpec(group_name, "-", co_cell=co)]

    cols: list[_ColumnSpec] = []
    for q in range(1, spec.questions + 1):
        cols.append(_ColumnSpec(group_name, f"Q{q}", co_cell=""))
    if spec.bonus_question:
        cols.append(
            _ColumnSpec(
                group_name,
                f"Bonus_Q{spec.questions + 1}",
                co_cell=None,
                is_bonus_question=True,
            )
        )
    cols.append(_ColumnSpec(group_name, "Total", co_cell=None, is_total=True))
    return cols


def _build_column_specs(components: list[MarksComponentSpec]) -> list[_ColumnSpec]:
    specs: list[_ColumnSpec] = []
    for group_name, comp in _resolve_component_names(components):
        specs.extend(_columns_for_component(group_name, comp))
    specs.append(_ColumnSpec("", "Result", co_cell=""))
    specs.append(_ColumnSpec("", "Grade_Point", co_cell=""))
    return specs


def _safe_filename_part(value: str) -> str:
    return "".join(c if c.isalnum() or c in "-_" else "_" for c in value.strip())


def build_constraint_workbook(body: MarksTemplateGenerateRequest) -> bytes:
    if not body.components:
        raise ValueError("Add at least one assessment component")

    columns = _build_column_specs(body.components)
    wb = Workbook()
    ws = wb.active
    ws.title = "Marks"

    # Columns A–B = Branch, Roll No.; assessments start at C (col 3)
    first_data_col = 3
    last_col = first_data_col + len(columns) - 1

    # --- Row 1: component group headers (merged per component) ---
    ws.cell(1, 1, "Branch")
    ws.cell(1, 2, "Roll No.")
    col_idx = first_data_col
    group_start: int | None = None
    current_group: str | None = None
    for spec in columns:
        if spec.group_name != current_group:
            if current_group and group_start is not None and col_idx > group_start:
                ws.merge_cells(
                    start_row=1,
                    start_column=group_start,
                    end_row=1,
                    end_column=col_idx - 1,
                )
            if spec.group_name:
                group_start = col_idx
                current_group = spec.group_name
                ws.cell(1, col_idx, spec.group_name)
            else:
                group_start = None
                current_group = spec.group_name
                if spec.subheader in ("Result", "Grade_Point"):
                    ws.cell(1, col_idx, spec.subheader)
        elif spec.group_name:
            ws.cell(1, col_idx, spec.group_name)
        col_idx += 1
    if current_group and group_start is not None and col_idx - 1 >= group_start:
        ws.merge_cells(
            start_row=1,
            start_column=group_start,
            end_row=1,
            end_column=col_idx - 1,
        )

    # Result / Grade_Point (no merge — single columns)
    for c in range(1, last_col + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = _HEADER_FILL

    # --- Row 2: subheaders ---
    ws.cell(2, 1, "Branch")
    ws.cell(2, 2, "Roll No.")
    for offset, spec in enumerate(columns):
        ws.cell(2, first_data_col + offset, spec.subheader)
        cell = ws.cell(2, first_data_col + offset)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = _HEADER_FILL

    # --- Row 3: CO row ---
    ws.cell(3, 1, "")
    ws.cell(3, 2, "CO to be entered")
    for offset, spec in enumerate(columns):
        value = spec.co_cell
        ws.cell(3, first_data_col + offset, value if value is not None else None)
        cell = ws.cell(3, first_data_col + offset)
        cell.font = Font(italic=True)
        cell.fill = _CO_FILL

    # --- Row 4: Max_Marks row ---
    ws.cell(4, 1, "")
    ws.cell(4, 2, "Max_Marks")
    for offset in range(len(columns)):
        ws.cell(4, first_data_col + offset, None)
    ws.cell(4, 2).font = Font(bold=True)

    # --- Rows 5–7: sample students ---
    for i, label in enumerate(("Student_1", "Student_2", "Student_3"), start=5):
        ws.cell(i, 1, "")
        ws.cell(i, 2, label)

    # Note row below samples
    note_row = 8
    ws.cell(
        note_row,
        1,
        "Delete sample rows (Student_1–Student_3) before uploading. Fill Branch, Roll No., marks, CO, and Max_Marks.",
    )
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=min(6, last_col))

    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 12
    for offset, spec in enumerate(columns):
        letter = get_column_letter(first_data_col + offset)
        if spec.subheader in ("Result", "Grade_Point"):
            ws.column_dimensions[letter].width = 12 if spec.subheader == "Grade_Point" else 10
        elif spec.subheader == "Total":
            ws.column_dimensions[letter].width = 10
        else:
            ws.column_dimensions[letter].width = 8

    ws.freeze_panes = "C5"

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def constraint_template_filename(course_code: str, semester: str) -> str:
    code = _safe_filename_part(course_code)
    sem = _safe_filename_part(semester)
    return f"{code}_{sem}_marks_template.xlsx"


@dataclass(frozen=True)
class AnalyzedQuestionSpec:
    label: str
    co_label: str
    max_marks: float
    is_bonus: bool = False


def build_analyzed_component_workbook(
    *,
    component_name: str,
    questions: list[AnalyzedQuestionSpec],
) -> bytes:
    """Single-component template with CO and Max_Marks pre-filled from question-paper analysis."""
    if not questions:
        raise ValueError("No questions to export.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Marks"

    group_name = component_name.strip() or "Component"
    first_data_col = 3
    columns: list[_ColumnSpec] = []
    for q in questions:
        sub = q.label if q.label else f"Q{len(columns) + 1}"
        if q.is_bonus:
            columns.append(_ColumnSpec(group_name, sub, co_cell=None, is_bonus_question=True))
        else:
            co = q.co_label if q.co_label else ""
            columns.append(_ColumnSpec(group_name, sub, co_cell=co))
    columns.append(_ColumnSpec(group_name, "Total", co_cell=None, is_total=True))
    columns.append(_ColumnSpec("", "Result", co_cell=""))
    columns.append(_ColumnSpec("", "Grade_Point", co_cell=""))

    last_col = first_data_col + len(columns) - 1

    ws.cell(1, 1, "Branch")
    ws.cell(1, 2, "Roll No.")
    ws.merge_cells(start_row=1, start_column=first_data_col, end_row=1, end_column=last_col - 2)
    ws.cell(1, first_data_col, group_name)
    ws.cell(1, last_col - 1, "Result")
    ws.cell(1, last_col, "Grade_Point")

    for c in range(1, last_col + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = _HEADER_FILL

    ws.cell(2, 1, "Branch")
    ws.cell(2, 2, "Roll No.")
    for offset, spec in enumerate(columns):
        ws.cell(2, first_data_col + offset, spec.subheader)
        cell = ws.cell(2, first_data_col + offset)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = _HEADER_FILL

    ws.cell(3, 1, "")
    ws.cell(3, 2, "CO")
    q_idx = 0
    for offset, spec in enumerate(columns):
        col = first_data_col + offset
        if spec.is_total or spec.subheader in ("Result", "Grade_Point"):
            ws.cell(3, col, None)
        else:
            q = questions[q_idx]
            ws.cell(3, col, None if q.is_bonus else (q.co_label or None))
            q_idx += 1
        cell = ws.cell(3, col)
        cell.font = Font(italic=True)
        cell.fill = _CO_FILL

    ws.cell(4, 1, "")
    ws.cell(4, 2, "Max_Marks")
    q_idx = 0
    for offset, spec in enumerate(columns):
        col = first_data_col + offset
        if spec.is_total:
            total = sum(q.max_marks for q in questions if not q.is_bonus)
            ws.cell(4, col, round(total, 2))
        elif spec.subheader in ("Result", "Grade_Point"):
            ws.cell(4, col, None)
        else:
            q = questions[q_idx]
            ws.cell(4, col, q.max_marks)
            q_idx += 1
    ws.cell(4, 2).font = Font(bold=True)

    for i, label in enumerate(("Student_1", "Student_2", "Student_3"), start=5):
        ws.cell(i, 1, "")
        ws.cell(i, 2, label)

    note_row = 8
    ws.cell(
        note_row,
        1,
        "Delete sample rows before uploading. Fill Branch, Roll No., and student marks per question.",
    )
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=min(8, last_col))

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 12
    for offset in range(len(columns)):
        letter = get_column_letter(first_data_col + offset)
        ws.column_dimensions[letter].width = 9

    ws.freeze_panes = "C5"

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
